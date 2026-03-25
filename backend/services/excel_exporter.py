import os
import datetime
import tempfile
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _thin_border():
    thin = Side(style="thin", color="D0D7E8")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _score_fill(final_score: float) -> PatternFill:
    if final_score >= 90:
        color = "1A7A4A"   # dark green
    elif final_score >= 75:
        color = "70AD47"   # light green
    elif final_score >= 60:
        color = "FFC000"   # amber
    elif final_score >= 45:
        color = "ED7D31"   # orange
    else:
        color = "C00000"   # red
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


HEADERS = [
    "#",
    "Candidate Details",
    "Years of\nExperience",
    "Current Company",
    "Education &\nCertifications",
    "Matched Skills",
    "Missing Skills",
    "Candidate Summary",
    "Technical Skills\nMatch",
    "Experience\nRelevance",
    "Domain\nKnowledge",
    "Role & Responsibility\nMatch",
    "Education\nScore",
    "Career Growth\n& Stability",
    "Final Weighted\nScore",
    "Tier",
]

COL_WIDTHS = [5, 36, 14, 22, 28, 28, 28, 42, 30, 30, 30, 30, 28, 30, 14, 24]


def export_to_excel(results: List[Dict], jd_filename: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Candidates"

    num_cols = len(HEADERS)

    # ── Title row ────────────────────────────────────────────────────────────
    jd_base = os.path.splitext(jd_filename)[0]
    date_str = datetime.datetime.now().strftime("%B %d, %Y · %H:%M")
    last_col = get_column_letter(num_cols)
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"RecruitIQ Screening Report  ·  {jd_base}  ·  {date_str}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0D1B3E", end_color="0D1B3E", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Header row ───────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="1A2952", end_color="1A2952", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[2].height = 36

    # ── Data rows ────────────────────────────────────────────────────────────
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=11, bold=True)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")

    for idx, result in enumerate(results):
        row_idx = idx + 3
        row_fill = alt_fill if idx % 2 == 1 else white_fill

        final_score = float(result.get("final_score", result.get("score", 0)) or 0)
        years = result.get("years_of_experience", "0") or "0"
        flag  = result.get("experience_flag", "")
        edu   = result.get("education", "N/A") or "N/A"
        edu_ok = result.get("education_meets_jd", False)

        # ── Col 1: Rank ──────────────────────────────────────────────────────
        c = ws.cell(row=row_idx, column=1, value=idx + 1)
        c.font = Font(name="Calibri", bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="top")

        # ── Col 2: Candidate Details ─────────────────────────────────────────
        name    = result.get("name", "N/A") or "N/A"
        phone   = result.get("phone", "N/A") or "N/A"
        email   = result.get("email", "N/A") or "N/A"
        linkedin = result.get("linkedin", "N/A") or "N/A"
        details_parts = [name, phone, email]
        if linkedin and linkedin != "N/A":
            details_parts.append(linkedin)
        details_value = "\n".join(details_parts)
        c2 = ws.cell(row=row_idx, column=2, value=details_value)
        c2.font = data_font
        # Make name bold via rich text isn't directly possible in openpyxl plain strings,
        # so we bold the entire cell and use font size to distinguish
        c2.font = Font(name="Calibri", size=10)
        if linkedin and linkedin != "N/A":
            c2.hyperlink = linkedin

        # ── Col 3: Years of Experience ────────────────────────────────────────
        exp_label = f"{years} year{'s' if years != '1' else ''} {flag}".strip()
        c3 = ws.cell(row=row_idx, column=3, value=exp_label)
        c3.font = Font(name="Calibri", size=11, bold=True)
        c3.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        # ── Col 4: Current Company ────────────────────────────────────────────
        company = result.get("current_company", result.get("latest_employment", "N/A")) or "N/A"
        c4 = ws.cell(row=row_idx, column=4, value=company)
        c4.font = data_font

        # ── Col 5: Education & Certifications ─────────────────────────────────
        edu_status = "✅ Meets JD requirement" if edu_ok else "❌ Does not meet JD"
        c5 = ws.cell(row=row_idx, column=5, value=f"{edu}\n{edu_status}")
        c5.font = data_font

        # ── Col 6: Matched Skills ─────────────────────────────────────────────
        matched = result.get("matched_skills", []) or []
        c6 = ws.cell(row=row_idx, column=6, value=", ".join(matched) if matched else "N/A")
        c6.font = data_font

        # ── Col 7: Missing Skills ─────────────────────────────────────────────
        missing = result.get("missing_skills", []) or []
        c7 = ws.cell(row=row_idx, column=7, value=", ".join(missing) if missing else "None")
        c7.font = data_font

        # ── Col 8: Candidate Summary ──────────────────────────────────────────
        c8 = ws.cell(row=row_idx, column=8, value=result.get("candidate_summary", "N/A") or "N/A")
        c8.font = data_font

        # ── Col 9–14: Score columns ───────────────────────────────────────────
        score_cols = [
            ("technical_skills_score", "technical_skills_reason"),
            ("experience_score",       "experience_reason"),
            ("domain_score",           "domain_reason"),
            ("role_score",             "role_reason"),
            ("education_score",        "education_reason"),
            ("career_score",           "career_reason"),
        ]
        for col_offset, (score_key, reason_key) in enumerate(score_cols):
            score_val  = result.get(score_key, 0) or 0
            reason_val = result.get(reason_key, "N/A") or "N/A"
            cell_val   = f"{score_val}/100 — {reason_val}"
            c = ws.cell(row=row_idx, column=9 + col_offset, value=cell_val)
            c.font = data_font

        # ── Col 15: Final Weighted Score ──────────────────────────────────────
        c15 = ws.cell(row=row_idx, column=15, value=round(final_score))
        c15.font = Font(name="Calibri", bold=True, size=14)
        c15.fill = _score_fill(final_score)
        c15.alignment = Alignment(horizontal="center", vertical="center")

        # ── Col 16: Tier ──────────────────────────────────────────────────────
        c16 = ws.cell(row=row_idx, column=16, value=result.get("tier", "N/A") or "N/A")
        c16.font = Font(name="Calibri", bold=True, size=10)
        c16.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ── Apply row background + borders ────────────────────────────────────
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            # Don't overwrite the score cell's special fill
            if col != 15:
                cell.fill = row_fill
            if col != 15:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _thin_border()

        ws.row_dimensions[row_idx].height = 80

    # ── Column widths ─────────────────────────────────────────────────────────
    for col, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A3"

    # ── Save ─────────────────────────────────────────────────────────────────
    date_tag = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
    safe_jd = "".join(c if c.isalnum() or c in "-_" else "_" for c in jd_base)
    filename = f"top_candidates_{safe_jd}_{date_tag}.xlsx"
    output_path = os.path.join(tempfile.gettempdir(), filename)
    wb.save(output_path)
    print(f"[EXCEL] Saved: {output_path}")
    return output_path
